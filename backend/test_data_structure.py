import openpyxl
import json

wb = openpyxl.load_workbook('test_files/simple_budget.xlsx', data_only=False)
ws = wb.active

# Simulate what our code does
cells = []

for row in ws.iter_rows():
    for cell in row:
        # Check if cell has content
        if cell.value is None:
            continue
        
        # Determine if cell has a formula
        has_formula = cell.data_type == 'f'
        formula_str = None
        
        if has_formula:
            # For formula cells, the value is the formula
            formula_str = f"={cell.value}" if not str(cell.value).startswith('=') else str(cell.value)
        
        cells.append({
            "row": cell.row,  # Keep 1-indexed
            "col": cell.column_letter,
            "value": str(cell.value) if cell.value else "",
            "formula": formula_str,
            "data_type": str(type(cell.value).__name__),
        })

sheets_data = [{
    "name": ws.title,
    "cells": cells,
    "row_count": ws.max_row,
    "col_count": ws.max_column,
}]

print(f"Total cells: {len(cells)}")
print(f"Cells with formulas: {sum(1 for c in cells if c['formula'])}")
print("\nFormula cells:")
for cell in cells:
    if cell['formula']:
        print(f"  {cell['col']}{cell['row']}: {cell['formula']}")

# Save to JSON for inspection
with open('test_sheets_data.json', 'w') as f:
    json.dump(sheets_data, f, indent=2)

print("\nSaved to test_sheets_data.json")
