"""
Analysis Service - Orchestrates the complete analysis workflow.
"""

import time
from typing import Dict, Any, Callable, Optional
from ..core import (
    FormulaParser,
    DAGBuilder,
    AnomalyDetector,
    CostDriverAnalyzer,
)
from ..utils import get_logger

logger = get_logger(__name__)


class AnalysisService:
    """
    Main service that orchestrates the complete analysis workflow.
    """
    
    def __init__(self):
        self.logger = logger.bind(component="AnalysisService")
    
    async def analyze_workbook(
        self,
        file_path: str,
        include_values: bool = False,
        detect_anomalies: bool = True,
        identify_cost_drivers: bool = True,
        top_drivers_count: int = 50,
        progress_callback: Optional[Callable[[int, str], None]] = None,
    ) -> Dict[str, Any]:
        """
        Perform complete workbook analysis.
        
        Args:
            file_path: Path to Excel file
            include_values: Include cell values in output
            detect_anomalies: Run anomaly detection
            identify_cost_drivers: Run cost driver analysis
            top_drivers_count: Number of top drivers to return
            progress_callback: Callback for progress updates
            
        Returns:
            Dictionary with analysis results
        """
        self.logger.info("Starting workbook analysis", file_path=file_path)
        
        try:
            # Step 1: Read Excel file (10-30%)
            if progress_callback:
                progress_callback(10, "Reading Excel file")
            
            sheets_data = await self._read_excel(file_path)
            
            if progress_callback:
                progress_callback(30, f"Read {len(sheets_data)} sheets")
            
            # Step 2: Build dependency graph (30-60%)
            if progress_callback:
                progress_callback(35, "Building dependency graph")
            
            dag_builder = DAGBuilder()
            graph = dag_builder.build_graph(sheets_data, include_values)
            
            if progress_callback:
                progress_callback(60, "Dependency graph complete")
            
            # Step 3: Detect anomalies (60-75%)
            anomalies_result = None
            if detect_anomalies:
                if progress_callback:
                    progress_callback(65, "Detecting anomalies")
                
                detector = AnomalyDetector(graph)
                anomalies = detector.detect_all(sheets_data)
                anomalies_result = detector.export_to_dict()
                
                if progress_callback:
                    progress_callback(75, f"Found {len(anomalies)} anomalies")
            
            # Step 4: Identify cost drivers (75-90%)
            cost_drivers_result = None
            if identify_cost_drivers:
                if progress_callback:
                    progress_callback(80, "Identifying cost drivers")
                
                analyzer = CostDriverAnalyzer(graph)
                drivers = analyzer.analyze(top_n=top_drivers_count)
                cost_drivers_result = analyzer.export_to_dict()
                
                if progress_callback:
                    progress_callback(90, f"Identified {len(drivers)} cost drivers")
            
            # Step 5: Prepare results (90-100%)
            if progress_callback:
                progress_callback(95, "Preparing results")
            
            # Calculate metrics
            graph_dict = dag_builder.export_to_dict()
            metrics = {
                "formula_count": sum(1 for node in graph_dict.get("nodes", []) if node.get("has_formula")),
                "input_count": sum(1 for node in graph_dict.get("nodes", []) if node.get("is_input")),
                "sheet_count": len(set(node.get("sheet") for node in graph_dict.get("nodes", []))),
                "avg_complexity": 0.0,  # Placeholder
            }
            
            result = {
                "graph": graph_dict,
                "anomalies": anomalies_result,
                "cost_drivers": cost_drivers_result,
                "metrics": metrics,
            }
            
            if progress_callback:
                progress_callback(100, "Analysis complete")
            
            self.logger.info("Workbook analysis complete")
            
            return result
            
        except Exception as e:
            self.logger.error("Analysis failed", error=str(e))
            raise
    
    async def _read_excel(self, file_path: str) -> list:
        """
        Read Excel file using Rust reader.
        
        Returns list of sheet dictionaries.
        """
        try:
            # Import Rust extension
            # import excel_reader
            # reader = excel_reader.ExcelReader(file_path)
            # reader.parse()
            
            # FORCE FALLBACK to openpyxl because Rust reader fails to extract formulas correctly (it reads values)
            raise ImportError("Forcing openpyxl fallback for correct formula extraction")
            
            sheets = reader.get_sheets()
            
            # Convert to dictionary format
            sheets_data = []
            for sheet in sheets:
                cells = []
                for cell in sheet.cells:
                    cells.append({
                        "row": cell.row,
                        "col": self._col_num_to_letter(cell.col),
                        "value": cell.value,
                        "formula": cell.formula,
                        "data_type": cell.data_type,
                    })
                
                sheets_data.append({
                    "name": sheet.name,
                    "cells": cells,
                    "row_count": sheet.row_count,
                    "col_count": sheet.col_count,
                })
            
            return sheets_data
            
        except ImportError:
            # Fallback to openpyxl if Rust extension not available
            self.logger.warning("Rust reader not available, using openpyxl")
            return await self._read_excel_fallback(file_path)
    
    async def _read_excel_fallback(self, file_path: str) -> list:
        """Fallback Excel reader using openpyxl or xlrd."""
        import os
        
        # Check file extension
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.xls':
            # Use xlrd for old .xls files
            try:
                import xlrd
                
                wb = xlrd.open_workbook(file_path)
                sheets_data = []
                
                for sheet_name in wb.sheet_names():
                    ws = wb.sheet_by_name(sheet_name)
                    cells = []
                    
                    for row_idx in range(ws.nrows):
                        for col_idx in range(ws.ncols):
                            cell = ws.cell(row_idx, col_idx)
                            
                            # Skip empty cells
                            if cell.ctype == xlrd.XL_CELL_EMPTY:
                                continue
                            
                            # Get cell value
                            value = cell.value
                            if cell.ctype == xlrd.XL_CELL_DATE:
                                value = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                            
                            cells.append({
                                "row": row_idx,
                                "col": self._col_num_to_letter(col_idx),
                                "value": str(value) if value else "",
                                "formula": None,  # xlrd doesn't support formulas
                                "data_type": self._get_xlrd_type(cell.ctype),
                            })
                    
                    sheets_data.append({
                        "name": sheet_name,
                        "cells": cells,
                        "row_count": ws.nrows,
                        "col_count": ws.ncols,
                    })
                
                return sheets_data
                
            except ImportError:
                raise Exception(
                    "xlrd library is required to read .xls files. "
                    "Please install it with: pip install xlrd"
                )
        else:
            # Use openpyxl for .xlsx, .xlsm files
            import openpyxl
            
            # Load workbook twice: once for formulas, once for values
            wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
            wb_values = openpyxl.load_workbook(file_path, data_only=True)
            sheets_data = []
            
            for sheet_name in wb_formulas.sheetnames:
                ws_formulas = wb_formulas[sheet_name]
                ws_values = wb_values[sheet_name]
                cells = []
                
                for row in ws_formulas.iter_rows():
                    for cell in row:
                        # Check if cell has content
                        if cell.value is None:
                            continue
                        
                        # Get corresponding value cell
                        value_cell = ws_values.cell(row=cell.row, column=cell.column)
                        
                        # Determine if cell has a formula
                        has_formula = cell.data_type == 'f'
                        formula_str = None
                        cell_value = None
                        
                        if has_formula:
                            # For formula cells, cell.value contains the formula string
                            formula_str = f"={cell.value}" if not str(cell.value).startswith('=') else str(cell.value)
                            # Get calculated value from data_only workbook
                            cell_value = value_cell.value
                        else:
                            # For non-formula cells, use the actual value
                            cell_value = cell.value
                        
                        cells.append({
                            "row": cell.row,  # Keep 1-indexed (Excel standard)
                            "col": cell.column_letter,
                            "value": str(cell_value) if cell_value is not None else "",
                            "formula": formula_str,
                            "data_type": str(type(cell.value).__name__),
                        })
                
                sheets_data.append({
                    "name": sheet_name,
                    "cells": cells,
                    "row_count": ws_formulas.max_row,
                    "col_count": ws_formulas.max_column,
                })
            
            return sheets_data
    
    def _get_xlrd_type(self, ctype: int) -> str:
        """Convert xlrd cell type to string."""
        import xlrd
        type_map = {
            xlrd.XL_CELL_EMPTY: "empty",
            xlrd.XL_CELL_TEXT: "str",
            xlrd.XL_CELL_NUMBER: "float",
            xlrd.XL_CELL_DATE: "datetime",
            xlrd.XL_CELL_BOOLEAN: "bool",
            xlrd.XL_CELL_ERROR: "error",
            xlrd.XL_CELL_BLANK: "blank",
        }
        return type_map.get(ctype, "unknown")
    
    def _col_num_to_letter(self, col_num: int) -> str:
        """Convert column number to letter (0 -> A, 25 -> Z, 26 -> AA)."""
        result = ""
        while col_num >= 0:
            result = chr(col_num % 26 + 65) + result
            col_num = col_num // 26 - 1
            if col_num < 0:
                break
        return result


# Create singleton instance
analysis_service = AnalysisService()
