"""
Generate a sample Excel workbook for testing Formula Intelligence
"""
import openpyxl
from openpyxl import Workbook
from datetime import datetime

# Create a new workbook
wb = Workbook()

# Sheet 1: Cost Drivers
ws1 = wb.active
ws1.title = "Cost_Drivers"

# Headers
ws1['A1'] = "Cost Item"
ws1['B1'] = "Unit Cost"
ws1['C1'] = "Quantity"
ws1['D1'] = "Total Cost"

# Data
cost_items = [
    ("Labor", 50, 100),
    ("Materials", 25, 200),
    ("Equipment", 100, 50),
    ("Overhead", 30, 150),
    ("Transportation", 15, 80)
]

row = 2
for item, unit_cost, qty in cost_items:
    ws1[f'A{row}'] = item
    ws1[f'B{row}'] = unit_cost
    ws1[f'C{row}'] = qty
    ws1[f'D{row}'] = f"=B{row}*C{row}"
    row += 1

# Total
ws1[f'A{row}'] = "TOTAL"
ws1[f'D{row}'] = f"=SUM(D2:D{row-1})"

# Sheet 2: Revenue
ws2 = wb.create_sheet("Revenue")
ws2['A1'] = "Product"
ws2['B1'] = "Price"
ws2['C1'] = "Units Sold"
ws2['D1'] = "Revenue"

products = [
    ("Product A", 100, 500),
    ("Product B", 150, 300),
    ("Product C", 200, 200),
]

row = 2
for product, price, units in products:
    ws2[f'A{row}'] = product
    ws2[f'B{row}'] = price
    ws2[f'C{row}'] = units
    ws2[f'D{row}'] = f"=B{row}*C{row}"
    row += 1

ws2[f'A{row}'] = "TOTAL REVENUE"
ws2[f'D{row}'] = f"=SUM(D2:D{row-1})"

# Sheet 3: Summary (with cross-sheet references)
ws3 = wb.create_sheet("Summary")
ws3['A1'] = "Financial Summary"
ws3['A3'] = "Total Revenue"
ws3['B3'] = "=Revenue!D5"

ws3['A4'] = "Total Costs"
ws3['B4'] = "=Cost_Drivers!D7"

ws3['A5'] = "Profit"
ws3['B5'] = "=B3-B4"

ws3['A7'] = "Profit Margin %"
ws3['B7'] = "=IF(B3>0, B5/B3*100, 0)"

ws3['A9'] = "Cost Breakdown"
ws3['A10'] = "Labor Cost"
ws3['B10'] = "=Cost_Drivers!D2"
ws3['A11'] = "Material Cost"
ws3['B11'] = "=Cost_Drivers!D3"
ws3['A12'] = "Equipment Cost"
ws3['B12'] = "=Cost_Drivers!D4"

# Sheet 4: Projections (with more complex formulas)
ws4 = wb.create_sheet("Projections")
ws4['A1'] = "Year"
ws4['B1'] = "Revenue Growth"
ws4['C1'] = "Projected Revenue"
ws4['D1'] = "Projected Costs"
ws4['E1'] = "Projected Profit"

years = [2024, 2025, 2026, 2027, 2028]
growth_rates = [0.10, 0.12, 0.15, 0.18, 0.20]

for idx, (year, growth) in enumerate(zip(years, growth_rates), start=2):
    ws4[f'A{idx}'] = year
    ws4[f'B{idx}'] = growth
    
    if idx == 2:
        # First year references Summary sheet
        ws4[f'C{idx}'] = "=Summary!B3"
        ws4[f'D{idx}'] = "=Summary!B4"
    else:
        # Subsequent years grow from previous year
        ws4[f'C{idx}'] = f"=C{idx-1}*(1+B{idx})"
        ws4[f'D{idx}'] = f"=D{idx-1}*1.08"  # 8% cost increase
    
    ws4[f'E{idx}'] = f"=C{idx}-D{idx}"

# Sheet 5: Analysis (with VLOOKUP, INDEX, MATCH)
ws5 = wb.create_sheet("Analysis")
ws5['A1'] = "Analysis Type"
ws5['B1'] = "Result"

ws5['A2'] = "Highest Cost Item"
ws5['B2'] = "=INDEX(Cost_Drivers!A:A, MATCH(MAX(Cost_Drivers!D:D), Cost_Drivers!D:D, 0))"

ws5['A3'] = "Lowest Revenue Product"
ws5['B3'] = "=INDEX(Revenue!A:A, MATCH(MIN(Revenue!D2:D4), Revenue!D2:D4, 0))"

ws5['A4'] = "Average Unit Cost"
ws5['B4'] = "=AVERAGE(Cost_Drivers!B2:B6)"

ws5['A5'] = "Total Units Sold"
ws5['B5'] = "=SUM(Revenue!C2:C4)"

ws5['A7'] = "Conditional Analysis"
ws5['A8'] = "High Margin Check"
ws5['B8'] = "=IF(Summary!B7>30, \"High Margin\", \"Low Margin\")"

# Add some intentional issues for anomaly detection
ws5['A10'] = "Test Error"
ws5['B10'] = "=InvalidRef!A1"  # Broken reference

ws5['A11'] = "Hardcoded Value"
ws5['B11'] = 12345  # Hard-coded value instead of formula

# Save the workbook
output_path = r"C:\Users\DELL\Desktop\landt\Formula-Intelligence\sample_workbook.xlsx"
wb.save(output_path)
print(f"Sample Excel workbook created successfully at: {output_path}")
print("\nWorkbook contains:")
print("- Cost_Drivers: Basic cost calculations")
print("- Revenue: Product revenue calculations")
print("- Summary: Cross-sheet references and summary metrics")
print("- Projections: Complex time-series projections")
print("- Analysis: Advanced formulas (INDEX, MATCH, VLOOKUP)")
print("\nFeatures included:")
print("✓ Multiple sheets with dependencies")
print("✓ Cross-sheet references")
print("✓ Complex formulas (SUM, IF, INDEX, MATCH, AVERAGE)")
print("✓ Intentional errors for anomaly detection")
print("✓ Hard-coded values for testing")
